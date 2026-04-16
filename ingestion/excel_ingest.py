import json
import os
import time
from typing import Literal, List, Dict, Any
from dotenv import load_dotenv
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from langchain_openai import ChatOpenAI

# ---------------------------------------------------------
# 1. SETUP & CONFIGURATION
# ---------------------------------------------------------
load_dotenv()

EXCEL_FILE = r"C:\Users\gauth\Downloads\260410 EXCEL GLORIA 2026_V60.xlsx"
OUTPUT_FILE = r"C:\Users\gauth\projects\ragbot-poc\data\ingested\excel_catalog_normalized.json"

# Ensure output directory exists
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

# ---------------------------------------------------------
# 2. PYDANTIC MODELS FOR STRUCTURED OUTPUT
# ---------------------------------------------------------
class ProductTaxonomy(BaseModel):
    row_id: str = Field(description="The unique ID passed in the prompt so we can map it back to the original row")
    animal: Literal["Dog", "Cat", "Bird", "Small Pet", "Fish", "Reptile", "Universal", "Unknown"] = Field(
        description="The primary animal this product is specifically for. If it doesn't apply to an animal, choose Unknown."
    )
    category: Literal["Food", "Toys", "Accessories", "Grooming & Hygiene", "Health", "Beds & Furniture", "Other"] = Field(
        description="The high-level category of the product."
    )
    sub_category: str = Field(description="A short descriptive sub-category (e.g. 'Collars', 'Nail Clippers')")

class BulkTaxonomyResponse(BaseModel):
    results: List[ProductTaxonomy]

# ---------------------------------------------------------
# 3. OPENROUTER LLM SETUP
# ---------------------------------------------------------
# Using Google's Gemini Flash 1.5 via OpenRouter for speed and cost-effectiveness
llm = ChatOpenAI(
    model="google/gemini-1.5-flash", 
    api_key=os.getenv("OPENROUTER_API_KEY"),
    base_url="https://openrouter.ai/api/v1",
    temperature=0
)

# Bind the Pydantic schema to force perfect JSON output
structured_llm = llm.with_structured_output(BulkTaxonomyResponse)


# ---------------------------------------------------------
# 4. PROCESSING LOGIC
# ---------------------------------------------------------
def process_batch(batch_list: List[Dict[str, Any]], brand: str) -> Dict[str, ProductTaxonomy]:
    """Sends a batch of products to the LLM and returns a dictionary mapped by unique reference ID."""
    
    # 1. Create a simplified prompt payload
    prompt_items = []
    for row in batch_list:
        name = row.get("PRODUCTO ES") or row.get("PRODUCTO") or row.get("PRODUCTO EN", "Unknown Product")
        ref_id = row.get("REFERENCIA") or row.get("REF. PRO.") or row.get("REF. PRO", "NO_ID")
        # Ensure ID is a string for stability
        ref_id = str(ref_id).strip()
        row["_internal_id"] = ref_id # Tag the original row with our guaranteed string ID
        
        prompt_items.append({"id": ref_id, "name": str(name).strip()})
        
    prompt = (
        f"Categorize these {len(batch_list)} pet products for the brand '{brand}'. "
        f"For each product, identify the target animal, the main category, and a sub-category. "
        f"You must return the results in JSON format mapped to their given IDs.\n\n"
        f"Products:\n{json.dumps(prompt_items, indent=2, ensure_ascii=False)}"
    )
    
    print(f"  [LLM] Requesting taxonomy for {len(batch_list)} products from brand '{brand}'...")
    
    # Retry logic just in case API fails
    max_retries = 3
    for attempt in range(max_retries):
        try:
            bulk_results = structured_llm.invoke(prompt)
            # Map results to a dictionary lookup by ID
            return {res.row_id: res for res in bulk_results.results}
        except Exception as e:
            print(f"  [LLM Error] Attempt {attempt+1} failed: {e}")
            if attempt == max_retries - 1:
                print("  [LLM Error] Skipping batch due to repeated failures.")
                return {}
            time.sleep(2)

def extract_schema(row_dict: dict, tab_name: str, taxonomy: ProductTaxonomy) -> dict:
    """Applies the schema mapping rules to collapse sparse columns."""
    
    # Core Fields
    sku = str(row_dict.get("REFERENCIA") or row_dict.get("REF. PRO.") or row_dict.get("REF. PRO", ""))
    ean = str(row_dict.get("COD EAN", ""))
    name_es = str(row_dict.get("PRODUCTO ES") or row_dict.get("PRODUCTO", ""))
    price = row_dict.get("PVPR") or row_dict.get("PVPR UNID") or row_dict.get("PVP UNITARIO")
    
    purchase_info = {}
    for p_key in ["CANTIDAD", "Unidad Mínima de Compra", "UNIDAD MÍNIMA DE VENTA", "UD. MIN.", "PEDIDO MÍNIMO"]:
        if row_dict.get(p_key):
            purchase_info[p_key] = row_dict[p_key]
            
    intl_names = {}
    for n_key in ["PRODUCTO EN", "PRODUCTO FR", "PRODUCTO PT", "PRODUCTO IT"]:
        if row_dict.get(n_key):
            intl_names[n_key.replace("PRODUCTO ", "")] = row_dict[n_key]
            
    # Soft Attributes (Grab everything else that has a value)
    ignore_keys = {"COD EAN", "PRODUCTO ES", "PRODUCTO", "REFERENCIA", "REF. PRO.", "REF. PRO", 
                   "PVPR", "PVPR UNID", "PVP UNITARIO", "ORDEN", "FOTO", "OBSERVACIONES", "ESTADO", "FECHA", "_internal_id"}
    ignore_keys.update(purchase_info.keys())
    ignore_keys.update(["PRODUCTO EN", "PRODUCTO FR", "PRODUCTO PT", "PRODUCTO IT"])
    
    attributes = {}
    for k, v in row_dict.items():
        if k not in ignore_keys and v is not None and str(v).strip() != "":
            attributes[k] = str(v).strip()

    metadata = {
        "origen_excel_row": row_dict.get("ORDEN"),
        "foto": row_dict.get("FOTO"),
        "observaciones": row_dict.get("OBSERVACIONES")
    }

    return {
        "brand": tab_name.strip(),
        "sku": sku.strip(),
        "ean": ean.strip(),
        "product_name_es": name_es.strip(),
        "product_names_intl": intl_names,
        "price": price,
        "purchase_info": purchase_info,
        "animal": taxonomy.animal if taxonomy else "Unknown",
        "category": taxonomy.category if taxonomy else "Other",
        "sub_category": taxonomy.sub_category if taxonomy else "Unknown",
        "attributes": attributes,
        "metadata": metadata
    }

# ---------------------------------------------------------
# 5. MAIN EXECUTION
# ---------------------------------------------------------
def main():
    print(f"Opening {EXCEL_FILE}...")
    wb = load_workbook(filename=EXCEL_FILE, read_only=True, data_only=True)
    
    all_normalized_products = []
    
    for sheet_name in wb.sheetnames:
        print(f"\nProcessing Tab: {sheet_name}")
        ws = wb[sheet_name]
        
        # Generator for reading the sheet
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
            
        columns = [str(c).strip() if c else f"U_{i}" for i, c in enumerate(header_row)]
        
        batch = []
        BATCH_SIZE = 50
        
        for row in rows_iter:
            # Skip completely empty rows
            if not any(row):
                continue
                
            row_dict = dict(zip(columns, row))
            batch.append(row_dict)
            
            if len(batch) >= BATCH_SIZE:
                taxonomy_map = process_batch(batch, sheet_name)
                for item in batch:
                    ref_id = item["_internal_id"]
                    tax_obj = taxonomy_map.get(ref_id)
                    normalized = extract_schema(item, sheet_name, tax_obj)
                    all_normalized_products.append(normalized)
                batch = [] # Reset batch
                
        # Process any remaining items in the last partial batch
        if batch:
            taxonomy_map = process_batch(batch, sheet_name)
            for item in batch:
                ref_id = item["_internal_id"]
                tax_obj = taxonomy_map.get(ref_id)
                normalized = extract_schema(item, sheet_name, tax_obj)
                all_normalized_products.append(normalized)

    # Output to massive JSON file
    print(f"\nWriting {len(all_normalized_products)} normalized records to {OUTPUT_FILE}...")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_normalized_products, f, indent=2, ensure_ascii=False)
        
    print("Done! 🎉 Your catalog is now clean, LLM-categorized, and ready to be dumped into Qdrant.")

if __name__ == "__main__":
    main()
