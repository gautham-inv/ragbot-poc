import json
from openpyxl import load_workbook
import sys

file_path = r"C:\Users\gauth\Downloads\260410 EXCEL GLORIA 2026_V60.xlsx"

try:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    results = {}
    all_columns_sets = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            cols = [str(c).strip() if c is not None else f"Unnamed_{i}" for i, c in enumerate(row)]
            break # only need the first row
            
        results[sheet_name] = cols
        all_columns_sets.append(set(cols))

    if all_columns_sets:
        common_cols = set.intersection(*all_columns_sets)
        union_cols = set.union(*all_columns_sets)
    else:
        common_cols = set()
        union_cols = set()

    output = {
        "status": "success",
        "tabs": wb.sheetnames,
        "columns_per_tab": results,
        "common_columns": list(common_cols),
        "all_unique_columns": list(union_cols)
    }

except Exception as e:
    output = {
        "status": "error",
        "message": str(e)
    }

with open("excel_output.json", "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2)

print("done")
